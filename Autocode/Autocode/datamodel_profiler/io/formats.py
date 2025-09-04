"""
Format-specific readers for different file types.
"""

import json
import yaml
from typing import Iterator, Union, Optional, Dict, Any, List
import pathlib
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

# Optional ORC support
try:
    import pyarrow.orc as orc
    ORC_AVAILABLE = True
except ImportError:
    orc = None
    ORC_AVAILABLE = False


class BaseReader:
    """Base class for format-specific readers."""
    
    def __init__(self, engine: str = "pandas"):
        self.engine = engine
    
    def read_file(
        self,
        path: pathlib.Path,
        chunksize: Optional[int] = None,
        **kwargs
    ) -> Iterator[pd.DataFrame]:
        """
        Read file and yield DataFrame chunks.
        
        Args:
            path: File path
            chunksize: Number of rows per chunk (None for single chunk)
            **kwargs: Format-specific options
            
        Yields:
            DataFrame chunks
        """
        raise NotImplementedError


class CSVReader(BaseReader):
    """CSV file reader."""
    
    def read_file(
        self,
        path: pathlib.Path,
        chunksize: Optional[int] = None,
        delimiter: str = ",",
        encoding: str = "utf-8",
        nrows: Optional[int] = None,
        **kwargs
    ) -> Iterator[pd.DataFrame]:
        """Read CSV file."""
        # Filter out non-CSV parameters
        csv_kwargs = {k: v for k, v in kwargs.items() 
                     if k not in ['json_lines', 'sheet_name']}
        
        read_kwargs = {
            "delimiter": delimiter,
            "encoding": encoding,
            "nrows": nrows,
            **csv_kwargs
        }
        
        if chunksize:
            read_kwargs["chunksize"] = chunksize
            for chunk in pd.read_csv(path, **read_kwargs):
                yield chunk
        else:
            yield pd.read_csv(path, **read_kwargs)


class ParquetReader(BaseReader):
    """Parquet file reader."""
    
    def read_file(
        self,
        path: pathlib.Path,
        chunksize: Optional[int] = None,
        columns: Optional[List[str]] = None,
        **kwargs
    ) -> Iterator[pd.DataFrame]:
        """Read Parquet file."""
        read_kwargs = {
            "columns": columns,
            **kwargs
        }
        
        if chunksize:
            # For parquet, we need to read in batches
            parquet_file = pq.ParquetFile(path)
            total_rows = parquet_file.metadata.num_rows
            
            for start in range(0, total_rows, chunksize):
                end = min(start + chunksize, total_rows)
                batch = parquet_file.read_row_group(
                    start // parquet_file.metadata.num_rows_per_row_group,
                    columns=columns
                )
                yield batch.to_pandas()
        else:
            yield pd.read_parquet(path, **read_kwargs)


class JSONReader(BaseReader):
    """JSON file reader."""
    
    def read_file(
        self,
        path: pathlib.Path,
        chunksize: Optional[int] = None,
        json_lines: bool = False,
        encoding: str = "utf-8",
        **kwargs
    ) -> Iterator[pd.DataFrame]:
        """Read JSON file."""
        with open(path, 'r', encoding=encoding) as f:
            if json_lines:
                # JSON Lines format
                data = []
                for line_num, line in enumerate(f):
                    if chunksize and line_num >= chunksize:
                        break
                    try:
                        data.append(json.loads(line.strip()))
                    except json.JSONDecodeError:
                        continue
                
                if data:
                    yield pd.DataFrame(data)
            else:
                # Regular JSON
                data = json.load(f)
                
                if isinstance(data, list):
                    # List of records
                    if chunksize:
                        for i in range(0, len(data), chunksize):
                            chunk_data = data[i:i + chunksize]
                            yield pd.DataFrame(chunk_data)
                    else:
                        yield pd.DataFrame(data)
                elif isinstance(data, dict):
                    # Single record or nested structure
                    yield pd.DataFrame([data])
                else:
                    raise ValueError("Unsupported JSON structure")


class YAMLReader(BaseReader):
    """YAML file reader."""
    
    def read_file(
        self,
        path: pathlib.Path,
        chunksize: Optional[int] = None,
        encoding: str = "utf-8",
        **kwargs
    ) -> Iterator[pd.DataFrame]:
        """Read YAML file."""
        with open(path, 'r', encoding=encoding) as f:
            data = yaml.safe_load(f)
            
            if isinstance(data, list):
                # List of records
                if chunksize:
                    for i in range(0, len(data), chunksize):
                        chunk_data = data[i:i + chunksize]
                        yield pd.DataFrame(chunk_data)
                else:
                    yield pd.DataFrame(data)
            elif isinstance(data, dict):
                # Single record
                yield pd.DataFrame([data])
            else:
                raise ValueError("Unsupported YAML structure")


class ExcelReader(BaseReader):
    """Excel file reader."""
    
    def read_file(
        self,
        path: pathlib.Path,
        chunksize: Optional[int] = None,
        sheet_name: Union[str, int, None] = 0,
        **kwargs
    ) -> Iterator[pd.DataFrame]:
        """Read Excel file."""
        read_kwargs = {
            "sheet_name": sheet_name,
            **kwargs
        }
        
        if chunksize:
            # Excel doesn't support chunking natively, so we read all and split
            df = pd.read_excel(path, **read_kwargs)
            for i in range(0, len(df), chunksize):
                yield df.iloc[i:i + chunksize]
        else:
            yield pd.read_excel(path, **read_kwargs)
    
    def get_sheet_names(self, path: pathlib.Path) -> List[str]:
        """Get list of sheet names in Excel file."""
        workbook = load_workbook(path, read_only=True)
        return workbook.sheetnames


class ORCReader(BaseReader):
    """ORC file reader."""
    
    def read_file(
        self,
        path: pathlib.Path,
        chunksize: Optional[int] = None,
        columns: Optional[List[str]] = None,
        **kwargs
    ) -> Iterator[pd.DataFrame]:
        """Read ORC file."""
        if not ORC_AVAILABLE:
            raise ImportError(
                "ORC support requires pyarrow>=10.0.0. "
                "Install with: pip install 'pyarrow>=10.0.0'"
            )
        
        try:
            # Try to read with pyarrow
            table = orc.read_table(path, columns=columns)
            
            if chunksize:
                # Convert to pandas and chunk
                df = table.to_pandas()
                for i in range(0, len(df), chunksize):
                    yield df.iloc[i:i + chunksize]
            else:
                yield table.to_pandas()
        except Exception as e:
            raise ValueError(f"Error reading ORC file: {e}")


class FormatDetector:
    """Utility class for detecting file formats."""
    
    @staticmethod
    def detect_from_content(path: pathlib.Path, sample_size: int = 1024) -> str:
        """
        Detect format from file content.
        
        Args:
            path: File path
            sample_size: Number of bytes to read for detection
            
        Returns:
            Detected format string
        """
        try:
            with open(path, 'rb') as f:
                sample = f.read(sample_size)
            
            # Check for common file signatures
            if sample.startswith(b'PK\x03\x04'):
                return "xlsx"
            elif sample.startswith(b'ORC'):
                return "orc"
            elif sample.startswith(b'PAR1'):
                return "parquet"
            elif sample.startswith(b'{') or sample.startswith(b'['):
                return "json"
            elif sample.startswith(b'---') or b'\n---\n' in sample:
                return "yaml"
            else:
                # Default to CSV
                return "csv"
        except Exception:
            return "csv"
    
    @staticmethod
    def is_binary_format(path: pathlib.Path) -> bool:
        """Check if file is in a binary format."""
        binary_extensions = {'.parquet', '.orc', '.xlsx', '.xls'}
        return path.suffix.lower() in binary_extensions
